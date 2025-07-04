import re
import pandas as pd
import streamlit as st
from io import BytesIO
from datetime import datetime, date
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import FILE_KEYWORDS, FIELD_MAPPINGS, pivot_config, RENAME_MAP
from excel_utils import (
    adjust_column_width, 
    highlight_replaced_names_in_main_sheet, 
    reorder_main_plan_by_unfulfilled_sheet, 
    format_currency_columns_rmb,
    format_thousands_separator,
    add_sheet_hyperlinks
)
from sheet_add import clean_df, append_all_standardized_sheets
from pivot_generator import generate_monthly_pivots, standardize_uploaded_keys
from file_utils import merge_cp_files_by_keyword, generate_fab_summary

class PivotProcessor:
    def process(self, uploaded_cp_files: dict, output_buffer):
        """
        替换品名、新建主计划表，并直接写入 Excel 文件（含列宽调整、标题行）。
        """
        # === 标准化上传文件名 ===
        self.cp_dataframes = {}
        cp_keywords = ["华虹", "先进", "DB", "上华1厂", "上华2厂", "上华5厂"]
        cp_file_counter = {k: 0 for k in cp_keywords}
        
        for filename, file_obj in uploaded_cp_files.items():
            matched = False
            for keyword in cp_keywords:
                if keyword in filename:
                    cp_file_counter[keyword] += 1
                    suffix = str(cp_file_counter[keyword])
                    new_key = f"{keyword}{suffix}" if cp_file_counter[keyword] > 1 else keyword
                    if keyword == "上华":
                        self.cp_dataframes[new_key] = pd.read_excel(file_obj, sheet_name = "wip")
                    else:
                        self.cp_dataframes[new_key] = pd.read_excel(file_obj)
                    matched = True
                    break
            if not matched:
                st.warning(f"⚠️ CP 文件 `{filename}` 未包含关键字，已跳过")

        self.cp_dataframes = merge_cp_files_by_keyword(self.cp_dataframes)

        st.write(self.cp_dataframes)

        df_fab_summary = generate_fab_summary(self.cp_dataframes)
         
        # === 写入 Excel 文件（主计划）===
        timestamp = datetime.now().strftime("%Y%m%d")
        with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
            # 写入主计划表
            df_fab_summary.to_excel(writer, sheet_name="FAB_WIP_汇总", index=False, startrow=1)
        
            # 获取 workbook 和 worksheet
            wb = writer.book
            ws = wb["FAB_WIP_汇总"]
        
            # 写时间戳和说明
            ws.cell(row=1, column=1, value=f"主计划生成时间：{timestamp}")            
    
            # 格式调整
            adjust_column_width(ws)

            # 设置字体加粗，行高也调高一点
            bold_font = Font(bold=True)
            ws.row_dimensions[2].height = 35
    
            # 遍历这一行所有已用到的列，对单元格字体加粗、居中、垂直居中
            max_col = ws.max_column
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.font = bold_font
                # 垂直水平居中
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 自动筛选
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A2:{last_col_letter}2"
        
            # 冻结
            ws.freeze_panes = "D3"
            append_all_standardized_sheets(writer, uploaded_files, self.additional_sheets)
           
        output_buffer.seek(0)
       
