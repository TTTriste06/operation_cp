import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import numbers
from openpyxl.styles import Font

def adjust_column_width(ws, max_width=70):
    """
    自动调整工作表中每列的宽度，适配内容长度。
    忽略第一行（通常用于填充非数据内容），以第二行起为基准。
    """
    for col_cells in ws.iter_cols(min_row=2):  # 从第二行开始
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 8, max_width)

def format_currency_columns_rmb(ws: Worksheet):
    """
    将所有标题中包含“金额”的列设置为人民币格式（¥#,##0.00），自动转换字符串为数字
    """
    header_row = 2
    max_col = ws.max_column
    max_row = ws.max_row

    for col_idx in range(1, max_col + 1):
        header = ws.cell(row=header_row, column=col_idx).value
        if isinstance(header, str) and "金额" in header:
            for row_idx in range(header_row + 1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                # 判断是否为数字或可转为数字的字符串
                if isinstance(val, (int, float)):
                    cell.number_format = u'¥#,##0.00'
                elif isinstance(val, str):
                    try:
                        num = float(val.replace(",", "").strip())
                        cell.value = num
                        cell.number_format = u'¥#,##0.00'
                    except:
                        pass  # 非法字符串忽略

def format_thousands_separator(ws: Worksheet):
    """
    将 C 列（第3列）之后的所有列，若为数值或可转数字的字符串，设置为千位分隔格式：#,##0.00。
    """
    header_row = 2
    max_col = ws.max_column
    max_row = ws.max_row

    for col_idx in range(3, max_col + 1):  # 从C列开始
        for row_idx in range(header_row + 1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value

            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
            elif isinstance(val, str):
                try:
                    num = float(val.replace(",", "").strip())
                    cell.value = num
                    cell.number_format = '#,##0'
                except:
                    continue

def add_sheet_hyperlinks(ws: Worksheet, sheet_names: list):
    """
    给 ws 的第二列添加超链接，跳转到名称相同的工作表。
    要求第 1 行是表头，从第 2 行开始为内容。

    参数：
        ws: openpyxl 的工作表对象（即“图”）
        sheet_names: 所有已存在的 sheet 名称列表
    """
    hyperlink_font = Font(color="0000FF", underline="single")  # 蓝色 + 下划线

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)
        target_sheet = cell.value
        if target_sheet and target_sheet in sheet_names:
            # 添加内部超链接
            cell.value = f'=HYPERLINK("#\'{target_sheet}\'!A1", "{target_sheet}")'
            cell.font = hyperlink_font  # 设置样式
